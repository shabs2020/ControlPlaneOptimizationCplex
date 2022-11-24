import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def create_workbook(fname):
    if os.path.isfile(fname):
        book = openpyxl.load_workbook(fname)
        book.remove(book["Links"])
        book.remove(book["Demands"])
        book.remove(book["Solution_Variables"])
    else:
        book = Workbook()

    return book


def write_link_details(book, links: dict):
    book.create_sheet("Links")
    sheet = book["Links"]
    sheet['A1'] = "No."
    sheet['B1'] = "Links_Ids"
    sheet['C1'] = "Link_Costs"
    sheet['D1'] = 'FirstEnd'
    sheet['E1'] = "SecondEnd"
    keys = [k for k in links.keys()]
    values = [v for v in links.values()]
    for l in range(len(links)):
        sheet['A' + str(l+2)] = l + 1
        sheet['B' + str(l+2)] = values[l][0]
        sheet['C' + str(l+2)] = values[l][1]
        sheet['D'+str(l+2)] = keys[l][0]
        sheet['E'+str(l+2)] = keys[l][1]
    return book


def write_demand_details(book, demands: dict):
    book.create_sheet("Demands")
    sheet = book["Demands"]
    sheet['A1'] = "No."
    sheet['B1'] = "Demand_Ids"
    sheet['C1'] = "Source"
    sheet['D1'] = "h(d)"
    sheet['E1'] = "Destination"
    sheet['F1'] = "Paths"
    sheet['G1'] = "Links in path"
    
    parent_row_num=2
    child_row_num=2
    i=1
    for d in demands:
        sheet['A' + str(parent_row_num)] = i
        sheet['B' + str(parent_row_num)] = demands[d][0]
        sheet['c' + str(parent_row_num)] = d
        sheet['D' + str(parent_row_num)] = demands[d][1]
        for p in demands[d][2]:
            sheet['E' + str(child_row_num)] = p[1]
            sheet['F' + str(child_row_num)] = p[0]
            ch=''
            for e in demands[d][2][p]:
                ch=ch+e+','
            ch = ch[:-1]
            sheet['G' + str(child_row_num)] = ch
            child_row_num=child_row_num+1
        
        
        #sheet.merge_cells('A2'+str(parent_row_num)+':'+'A'+ str(parent_row_num+len(demands[d][2])-1))
        parent_row_num = parent_row_num +  len(demands[d][2]) 
        child_row_num = parent_row_num
        i+=1

    return book


def write_solution(book, dataframe):
    book.create_sheet("Solution_Variables")
    sheet=book["Solution_Variables"]
    for r in dataframe_to_rows(dataframe, index=True, header=True):
        sheet.append(r)
    return book


def save_book(book: Workbook, fname: str):
    book.save(fname)
