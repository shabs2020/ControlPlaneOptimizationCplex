import networkx as nx
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pandas as pd 
import numpy as np
import math
import cplex_input
file_path = os.path.abspath(os.path.join(__file__, "../.."))
BASE_DIR = os.path.dirname(file_path)

def calc_nodecapacity(fname,s_factor):

    demands_sheet=pd.read_excel(BASE_DIR + '/'+ fname, sheet_name='Demands',header=0)
    nodes=[]
    direct_nodes=[i for i in demands_sheet['Destination'].unique() ]
    nodes.extend(direct_nodes)
    nodes.extend(i for  i in demands_sheet['Source'].unique())
    nodes=[i for i in nodes if pd.notnull(i)]
    links_sheet=pd.read_excel(BASE_DIR + '/'+ fname, sheet_name='Links',header=0)
    variables_in_sol=pd.read_excel(BASE_DIR + '/'+ fname, sheet_name='Solution_Variables',header=0)
    variables_in_sol=variables_in_sol.dropna(axis=0)
    print("direct nodes are {}".format(direct_nodes))
    path_per_source={}
    for d in variables_in_sol["name"]:
        s_name = d[4:-3].translate({ord("_"): None})
        path_name = d[-3:].translate({ord("_"): None})
        if s_name not in list(path_per_source.keys()):
            path_per_source[s_name]=[path_name]
        else:
            path_per_source[s_name].append(path_name)
        

        # capacity[demand_paths["d_" + s_name][path_name][-1]] = (
        #     demand_volume["d_" + s_name] / 2
        #     + network.nodes[demand_paths["d_" + s_name][path_name][-1]][
        #         "demandVolume"
        #     ]
        #     if demand_paths["d_" + s_name][path_name][-1]
        #     not in capacity.keys()
        #     else capacity[demand_paths["d_" + s_name][path_name][-1]]
        #         + demand_volume["d_" + s_name] / 2
    network=nx.Graph()
    for n in nodes:
        network.add_node(n, id=nodes.index(n))

    for e in range(len(links_sheet['No.'])):
        network.add_edge(links_sheet['FirstEnd'][e], links_sheet['SecondEnd'][e], id=links_sheet['Links_Ids'][e])

    path_index_per_demand= demands_sheet[~demands_sheet['Source'].isnull()].index.tolist()
    path_index_per_demand.append(len(demands_sheet))

    routes=[]
    demand_per_control_node={}
    for src in path_per_source:

        for elem,next_elem in zip(path_index_per_demand, path_index_per_demand[1:]):

            if demands_sheet['Source'][elem]==src:

                demand=[demands_sheet['h(d)'][i] for i in range(elem, next_elem) if not pd.isna(demands_sheet['h(d)'][i])]

                for i in range(elem, next_elem):
                
                    if demands_sheet['Path Ids'][i] in path_per_source[src]:
                        destination=demands_sheet['Destination'][i]
                        #print('Path{} has destination{}'.format(demands_sheet['Path Ids'][i], destination))
                        if destination in list(demand_per_control_node.keys()):
                            demand_per_control_node[destination]+=demand[0]/2
                        else:
                            demand_per_control_node[destination]=demand[0]/2
                            #print("demand {}, with path {} has destination {}".format(src,demands_sheet['Path Ids'][i],destination))
                    # routes.append(demands_sheet['Paths'][i].split("-"))
                    # paths[(demands_sheet['Path Ids'][i])]=demands_sheet['Paths'][i].split("-")

                # for i in range(elem, next_elem):
                #     routes.append(demands_sheet['Paths'][i].split("-"))
                    #paths[(demands_sheet['Path Ids'][i])]=demands_sheet['Paths'][i].split("-") 
    #print(demand_per_control_node)
    for n in demand_per_control_node:
        demand_on_node=1.86 + 1.18 + (0.86 + 0.745 + 0.2 + 0.1) * network.degree(n)
        demand_per_control_node[n]+=demand_on_node
        demand_per_control_node[n]=cplex_input.round_capacity(demand_per_control_node[n])
   
    for d in direct_nodes:
        if d not in list(demand_per_control_node.keys()):
            demand_per_control_node[d] =cplex_input.round_capacity((1.86 + 1.18 + (0.86 + 0.745 + 0.2 + 0.1) * network.degree(d)*s_factor))
    d=0
    for n in network.nodes:
        #s=cplex_input.round_capacity(1.86 + 1.18 + (0.86 + 0.745 + 0.2 + 0.1) * network.degree(n))*s_factor)
        d+=cplex_input.round_capacity((1.86 + 1.18 + (0.86 + 0.745 + 0.2 + 0.1) * network.degree(n))*s_factor)
    print(d)

    #print(demand_per_control_node)
    #print(len(demand_per_control_node))
    return demand_per_control_node


fname= 'Topologies/Coronet30.xlsx'
df_nodes = pd.read_excel(BASE_DIR + '/'+ fname, sheet_name="Nodes")
df_links = pd.read_excel(BASE_DIR + '/'+ fname, sheet_name="Links")
graph = nx.Graph()
for n in range(0, len(df_nodes)):
    graph.add_node(
        df_nodes["Name"][n],
        lon=df_nodes["Longitude"][n],
        lat=df_nodes["Latitude"][n],
        pos=(df_nodes["Longitude"][n], df_nodes["Latitude"][n]),
    )
for e in range(0, len(df_links)):
    graph.add_edge(
        df_links["Node-A"][e],
        df_links["Node-Z"][e],
        linkDist=round(df_links["Length"][e], 3),
    )
for e in graph.edges:
    link_cost = 1
    graph[e[0]][e[1]]["linkCost"] = link_cost
demand=0
for n in graph.nodes:
    demand+= 1.86 + 1.18 + (0.86 + 0.745 + 0.2 + 0.1) * graph.degree(n)
avg_demand=demand/len(graph.nodes)
print(avg_demand)
print(cplex_input.round_capacity(avg_demand))

# node_cost={}

# for i in range(2,17):
#     fname='Stats/Euclid/Model_Stats_G17_M'+ str(i) +'_1.xlsx'
#     demand_per_control_node=calc_nodecapacity(fname,1)
#     node_cost[i]=sum(demand_per_control_node.values())
#     book=openpyxl.load_workbook(BASE_DIR+'/'+'Stats/NewFormulation/NewFormulation/Objectives_NewFormTest1.xlsx')
#     sheet=book['Obj_Values']
#     sheet['E'+str(i)]=sum(demand_per_control_node.values())
#     book.save(BASE_DIR+'/'+'Stats/NewFormulation/NewFormulation/Objectives_NewFormTest1.xlsx')
# print(node_cost)



# fname='Stats/Euclid_New/Model_Stats_G17_Scaled_M16_10.xlsx'
# node_cost={}
# j=2
# for i in range(2,17):
#     fname='Stats/Euclid_New/Model_Stats_G17_Scaled_M2_1.xlsx'
#     demand_per_control_node=calc_nodecapacity(fname,1)
#     node_cost[i]=sum(demand_per_control_node.values())
#     book=openpyxl.load_workbook(BASE_DIR+'/'+'Stats/Coronet30/Objectives.xlsx')
#     sheet=book['Obj_Values']
#     sheet['E'+str(j)]=sum(demand_per_control_node.values())
#     book.save(BASE_DIR+'/'+'Stats/Coronet30/Objectives.xlsx')
#     j+=1

# print(node_cost)

# print(len(demand_per_control_node))
# node_cost={}
# fname1='Stats/Euclid_New/Model_Stats_G17_Scaled_M2_10.xlsx'
# for i in range(2,17):
#     fname='Stats/Euclid_New/Model_Stats_G17_Scaled_M'+ str(i) +'_10.xlsx'
#     demand_per_control_node=calc_nodecapacity(fname,10)
#     node_cost[i]=sum(demand_per_control_node.values())
#     book=openpyxl.load_workbook(BASE_DIR+'/'+'Stats/Euclid_New/Objectives_Scaled.xlsx')
#     sheet=book['PL_NodeCosts']
#     sheet['E'+str(i)]=sum(demand_per_control_node.values())
#     book.save(BASE_DIR+'/'+'Stats/Euclid_New/Objectives_Scaled.xlsx')

# print(node_cost)

