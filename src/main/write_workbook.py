import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def create_workbook(fname):
    if os.path.isfile(fname):
        book = openpyxl.load_workbook(fname)
        sheet_names = book.sheetnames
        for i in sheet_names:
            del book[i]
    else:
        book = Workbook()

    return book


def write_link_details(book, links: dict):
    book.create_sheet("Links")
    sheet = book["Links"]
    sheet['A1'] = "No."
    sheet['B1'] = "Links_Ids"
    sheet['C1'] = "Link_Costs"
    sheet['D1'] = 'FirstEnd'
    sheet['E1'] = "SecondEnd"
    keys = [k for k in links.keys()]
    values = [v for v in links.values()]
    for l in range(len(links)):
        sheet['A' + str(l+2)] = l + 1
        sheet['B' + str(l+2)] = values[l][0]
        sheet['C' + str(l+2)] = values[l][1]
        sheet['D'+str(l+2)] = keys[l][0]
        sheet['E'+str(l+2)] = keys[l][1]
    return book


def write_demand_details(book, demand_volume: dict, demand_paths: dict, demand_path_edges: dict, demand_path_lengths: dict):
    book.create_sheet("Demands")
    sheet = book["Demands"]
    sheet['A1'] = "No."
    sheet['B1'] = "Demand_Ids"
    sheet['C1'] = "Source"
    sheet['D1'] = "h(d)"
    sheet['E1'] = "Destination"
    sheet['F1'] = "Path Ids"
    sheet['G1'] = "Paths"
    sheet['H1'] = "Links in path"
    sheet['I1'] = "Path Lengths"

    parent_row_num = 2
    child_row_num = 2
    for d in demand_volume:
        sheet['A' + str(child_row_num)] = parent_row_num
        sheet['B' + str(child_row_num)] = d
        sheet['c' + str(child_row_num)] = d[2:]
        sheet['D' + str(child_row_num)] = demand_volume[d]

        for p in demand_paths[d]:
            sheet['E' + str(child_row_num)] = demand_paths[d][p][-1]
            sheet['F' + str(child_row_num)] = p
            sheet['G' + str(child_row_num)
                  ] = '-'.join(s for s in demand_paths[d][p])
            sheet['H' + str(child_row_num)
                  ] = '-'.join(s for s in demand_path_edges[d][p])
            sheet['I' + str(child_row_num)] = demand_path_lengths[d][p]
            child_row_num += 1
        #sheet.merge_cells('A2'+str(parent_row_num)+':'+'A'+ str(parent_row_num+len(demands[d][2])-1))
        parent_row_num +=1
    return book


def write_solution(book, dataframe):
    book.create_sheet("Solution_Variables")
    sheet = book["Solution_Variables"]
    for r in dataframe_to_rows(dataframe, index=True, header=True):
        sheet.append(r)
    return book

def write_kpis_per_M(book,min_objective_value, link_kpi, path_kpi,control_node_costs):
    book.create_sheet("Objective_Values")
    sheet = book["Objective_Values"]
    sheet['A1'] = "Obj.Value."
    sheet['B1'] = "LinkBW"
    sheet['C1'] = "PathUtils"
    sheet['D1'] = "NodeCosts"
    sheet['A2'] = min_objective_value
    sheet['B2'] = link_kpi
    sheet['C2'] = path_kpi
    sheet['D2'] = control_node_costs

    return book

def write_objective_values(book, obj: dict,kpi1_perf:dict, kpi2_perf:dict):
    book.create_sheet("Obj_Values")
    sheet = book["Obj_Values"]
    sheet['A1'] = "Nodes Num."
    sheet['B1'] = "Obj. Value"
    sheet['C1'] = "LinkBW"
    sheet['D1'] = "PathUtils"
    sheet['E1'] = "NodeCosts"
    i = 2
    for val in obj:
        sheet['A' + str(i)] = val
        sheet['B' + str(i)] = obj[val]
        if kpi1_perf.get(val) is not None:
            sheet['C' + str(i)] = kpi1_perf[val]
        else:
            sheet['C' + str(i)] = 0
        if kpi2_perf.get(val) is not None:
            sheet['D' + str(i)] = kpi2_perf[val][0]
            sheet['E' + str(i)] = kpi2_perf[val][1]
        else:
            sheet['D' + str(i)] = 0
        
        i += 1
    return book
def write_nested_dict(sheet, obj:dict):
    sheet['A1'] = 'Num_Nodes'
    keys=list(obj.keys())
    n_nodes=[k for k in obj[1]]
    val_1=[k[1] for k in obj[1].items()]
    val_10=[k[1] for k in obj[10].items()]
    val_100=[k[1] for k in obj[100].items()]
    val_1000=[k[1] for k in obj[1000].items()]
    val_5000=[k[1] for k in obj[5000].items()]
    
    if(type(val_1[0]) is list):
        k=0
        for i in range(2, len(obj)+2):
            j=0
            while j<2:
                sheet.cell(row=1, column=i+k,value=keys[i-2]) 
                k+=1
                j+=1
            k-=1
        for i in range(0,len(n_nodes)):
            sheet['A' + str(i+2)] = n_nodes[i]
            sheet['B' + str(i+2)] =val_1[i][0]
            sheet['C' + str(i+2)] =val_1[i][1]
            sheet['D' + str(i+2)] =val_10[i][0]
            sheet['E' + str(i+2)] =val_10[i][1]
            sheet['F' + str(i+2)] =val_100[i][0]
            sheet['G' + str(i+2)] =val_100[i][1]
            sheet['H' + str(i+2)] =val_1000[i][0]
            sheet['I' + str(i+2)] =val_1000[i][1]
            sheet['J' + str(i+2)] =val_5000[i][0]
            sheet['K' + str(i+2)] =val_5000[i][1]

            #print(i)
    else:
        for i in range(2, len(obj)+2):
            sheet.cell(row=1, column=i,value=keys[i-2]) 
        for i in range(0,len(n_nodes)):
            sheet['A' + str(i+2)] = n_nodes[i]
            sheet['B' + str(i+2)] =val_1[i]
            sheet['C' + str(i+2)] =val_10[i]
            sheet['D' + str(i+2)] =val_100[i]
            sheet['E' + str(i+2)] =val_1000[i]
            sheet['F' + str(i+2)] =val_5000[i]

    return sheet

def write_objective_values_scaled(book, obj: dict, sheet_name:str):
    book.create_sheet(sheet_name)
    sheet = book[sheet_name]
   
    sheet=write_nested_dict(sheet, obj)
    return book

def save_book(book: Workbook, fname: str):
    book.save(fname)


def load_workbook(f_name, sheet_name):
    cell_values = []
    book = openpyxl.load_workbook(f_name)
    sheet1 = book[sheet_name]
    for col in sheet1.columns:
        cell_values.append([data.value for data in col[1:]])
    return cell_values
